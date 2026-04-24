"""
Copyright (c) 2026 건원건축 김정현. All rights reserved.

본 프로그램은 도면 검토 업무 효율화를 위해 기획 및 개발되었습니다.
외부 업체로의 유출, 무단 복제 및 소스코드 수정을 엄격히 금지합니다.


app.py  —  DWG 자동 검토기 v_1.7 Ultimate Edition (Kunwon Masterpiece)
========================================================================
[V6.7 업데이트]
1. Drag & Drop 완벽 지원: 윈도우 탐색기에서 파일(.dwg)이나 폴더를 마우스로 끌어서 
   프로그램 창에 던지면(Drop) 경로가 자동으로 인식되고 세팅됩니다.
2. 스마트 숨김형 UI (Progressive Disclosure): 평소에는 심플하게 1개의 도곽 이름만 받지만, 
   체크박스를 켜면 [목록표(Master) / 개별도면(Slave)] 도곽 이름을 분리해서 탐색합니다.
   (박스 좌표는 무조건 Master 기준, 개별도면 탐색은 Slave 이름 기준으로 작동)
3. 무한 체인 정규식 (V6.6 로직 유지): AA-000-000-000 무제한 추출 완벽 적용.
========================================================================
"""

from __future__ import annotations
import glob, os, re, sys, webbrowser, json, math, logging
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import concurrent.futures
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import ezdxf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import customtkinter as ctk
from tkinterdnd2 import TkinterDnD, DND_FILES  # [V6.7 추가] 드래그 앤 드롭 엔진

# ============================================================================
# [UI 테마 설정]
# ============================================================================
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# ============================================================================
# [로깅 설정] GUI 핸들러는 앱 초기화 시 추가, 파일 핸들러는 즉시 활성화
# ============================================================================
logger = logging.getLogger("AutoDWG")
logger.setLevel(logging.DEBUG)
logger.addHandler(logging.NullHandler())

def _setup_file_logger():
    log_dir = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'AutoDWG_Checker')
    os.makedirs(log_dir, exist_ok=True)
    fh = logging.FileHandler(os.path.join(log_dir, 'autodwg.log'), encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)-8s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    logger.addHandler(fh)

_setup_file_logger()

리포트_이름: str = "도면검토리포트_최종.xlsx"
ODA_DOWNLOAD_URL = "https://www.opendesign.com/guestfiles/oda_file_converter"

# ============================================================================
# 0. 기존 코어 엔진 (JSON 로드 및 ODA)
# ============================================================================
def load_roi_config(block_name: str) -> Optional[dict]:
    config_dir = os.path.join(os.environ.get('APPDATA', ''), 'AutoDWG_Checker')
    config_path = os.path.join(config_dir, f"{block_name}.json")
    if os.path.exists(config_path):
        for enc in ['cp949', 'utf-8', 'euc-kr']:
            try:
                with open(config_path, 'r', encoding=enc) as f:
                    return json.load(f)
            except Exception: continue
    return None

def _oda_환경_설정() -> str:
    found_path = ""
    for 경로 in [r"C:\Program Files\ODA", r"C:\Program Files (x86)\ODA"]:
        실행파일들 = glob.glob(os.path.join(경로, "**", "ODAFileConverter.exe"), recursive=True)
        if 실행파일들:
            found_path = sorted(실행파일들, reverse=True)[0]
            break
    if found_path:
        폴더경로 = os.path.dirname(found_path)
        if 폴더경로 not in os.environ.get("PATH", ""):
            os.environ["PATH"] = 폴더경로 + os.pathsep + os.environ.get("PATH", "")
        try:
            ezdxf.options.odafc_win_exec_path = found_path
        except AttributeError:
            if not ezdxf.options.has_section('odafc'):
                ezdxf.options._config.add_section('odafc')
            ezdxf.options.set('odafc', 'win_exec_path', found_path)
    return found_path

# ============================================================================
# 1. 공통 유틸리티 (무한체인 정규식 및 필터)
# ============================================================================
_도면번호_패턴 = re.compile(r"(?<![가-힣A-Za-z0-9])([A-Z\u0391-\u03A9\.가-힣][A-Z0-9\u0391-\u03A9\.가-힣]{0,4})[\s\-_~–—−]*(\d{1,5}(?:[\s\-_~–—−]+\d{1,5}(?![가-힣㎡,]))*[A-Za-z]*|TOE)(?!\d|[A-Za-z]|[가-힣])")
_축척_패턴 = re.compile(r"(1\s?[/:,]\s?([\d,]+)|NONE|N/A)", re.I)
_동_패턴 = re.compile(r"((?:[0-9A-Za-z]+\s*[,~&]\s*)*[0-9A-Za-z]+동)")
_동_제외단어 = ["인동", "주동", "공동", "자동", "수동", "전동", "연동", "이동", "작동", "부동", "진동", "명동", "구동", "개동", "각동", "해당동", "상동", "하동"]

GLOBAL_IGNORE_HEADERS = [
    "SUBJECT TITLE", "SUBJECT", "PROJECT TITLE", "PROJECT",
    "DRAWING TITLE", "DRAWING NO.", "DRAWING NO", "DWG.NO.", "DWG. NO.", "DWG.NO", "DWG NO.", "DWG NO", "TITLE",
    "SHEET NO.", "SHEET NO", "SHT NO.", "SHT NO", "SHEET",
    "도면번호", "도연번호", "일련번호", "연번", "NO", "NO.", "도면명", "도면명칭", "축척(A1)", "축척(A3)", "축척(A0)", 
    "SCALE(A1)", "SCALE(A3)", "SCALE(A0)", "축척(1:)", "축척(1/)", "SCALE(1:)", "SCALE(1/)", "(1:)", "(1/)",
    "축척", "축적", "SCALE", "비고", "REMARK", "REMARKS", "사업승인", "착공", "견적", "사용승인", "1:1"
]
CATEGORY_KEYWORDS = ["공통사항", "일반사항", "건축도면", "구조도면", "기계도면", "전기도면", "토목도면", "조경도면", "소방도면", "부분상세도"]

def _clean_text_from_headers(txt: str) -> str:
    clean = txt
    for h in sorted(GLOBAL_IGNORE_HEADERS, key=len, reverse=True):
        clean = re.compile(re.escape(h), re.IGNORECASE).sub(" ", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    return re.sub(r"^[-_,\s]+|[-_,\s]+$", "", clean)

def _extract_dong_from_title(title: str) -> str:
    matches = list(_동_패턴.finditer(title))
    for m in matches:
        dong_str = m.group(1)
        if not any(ex_word in dong_str for ex_word in _동_제외단어): return dong_str
    return ""

def _extract_group_prefix(title: str) -> str:
    """Return the part of title that precedes any floor/content keyword (e.g. '코어#1' from '코어#1 지하6층 평면도')."""
    m = re.search(r"지[하상]\d*층|B\d+F?|\d+층|옥상|기초|파일", title)
    if m and m.start() > 0:
        return re.sub(r"[,\s]+$", "", title[:m.start()])
    return ""

def _도면번호_세척(raw_s: str) -> str:
    if not raw_s: return ""
    suffix_m = re.search(r"[a-z]+$", raw_s.strip())
    orig_suffix = suffix_m.group(0) if suffix_m else ""
    s = raw_s.strip().upper().replace("Λ", "A").replace("Δ", "A").replace("TOE", "108")
    if s.startswith("."): s = "AA" + s[1:]
    s = re.sub(r"\s*([-_~])\s*", r"\1", s)
    s = re.sub(r"[-_~]{2,}", "-", re.sub(r"\s+", " ", s))
    # CAD에서 한 글자씩 분리 저장된 경우 합치기 (예: "AA-0 0 0-0 0 0-0" -> "AA-000-000-0")
    segs = re.split(r"([-_~])", s)
    merged = []
    for i, seg in enumerate(segs):
        if i % 2 == 1:
            merged.append(seg)
        else:
            parts = [t for t in seg.split(" ") if t]
            if parts and all(len(t) == 1 for t in parts):
                merged.append("".join(parts))
            elif parts:
                buf, rp = [], []
                for p in parts:
                    if len(p) == 1: buf.append(p)
                    else:
                        if buf: rp.append("".join(buf)); buf = []
                        rp.append(p)
                if buf: rp.append("".join(buf))
                merged.append(" ".join(rp))
            else:
                merged.append(seg)
    s = "".join(merged)
    # 남은 공백은 그래픽 대시가 텍스트로 저장되지 않은 자리 -> 대시로 교체
    s = re.sub(r"(?<=[A-Za-z0-9]) (?=[A-Za-z0-9])", "-", s)
    if orig_suffix: s = s[:-len(orig_suffix)] + orig_suffix
    return s

def _spatial_reconstruct_num_str(texts: list) -> str:
    """Join number-column text entities; insert '-' where spatial gap between
    consecutive single-char alphanumeric tokens indicates a missing graphical dash."""
    if not texts: return ""
    single_char_gaps = []
    for i in range(1, len(texts)):
        tx, _, txt_i, _ = texts[i]
        px, _, ptxt_i, _ = texts[i-1]
        s, ps = txt_i.strip(), ptxt_i.strip()
        if (len(s) == 1 and len(ps) == 1
                and re.match(r"[0-9A-Za-z]", s)
                and re.match(r"[0-9A-Za-z]", ps)):
            single_char_gaps.append(tx - px)
    if len(single_char_gaps) >= 2:
        sorted_gaps = sorted(single_char_gaps)
        median_gap = sorted_gaps[len(sorted_gaps) // 2]
        gap_threshold = median_gap * 1.6
    elif single_char_gaps:
        avg_h = sum(t[3] for t in texts) / len(texts)
        gap_threshold = avg_h * 0.85
    else:
        gap_threshold = None
    tokens = []
    for i, t in enumerate(texts):
        tx, ty, txt_i, th = t
        stripped = txt_i.strip()
        if not stripped:
            continue
        if i > 0 and gap_threshold is not None:
            prev_tx, _, prev_txt_i, _ = texts[i-1]
            ps = prev_txt_i.strip()
            if (len(ps) == 1 and len(stripped) == 1
                    and re.match(r"[0-9A-Za-z]", ps)
                    and re.match(r"[0-9A-Za-z]", stripped)
                    and (tx - prev_tx) > gap_threshold):
                tokens.append("-")
        tokens.append(stripped)
    return " ".join(tokens)

def _merge_title_char_runs(s: str) -> str:
    """Merge space-separated single-char tokens in title strings (char-by-char CAD storage)."""
    if not s: return ""
    result_parts = []
    run = []
    for tok in s.split(" "):
        if tok and len(tok) == 1:
            run.append(tok)
        else:
            if run:
                merged = "".join(run)
                # Attach separator-leading runs directly to previous word (e.g. "-12" -> "근거-12")
                if result_parts and (merged[0] in "-_~" or result_parts[-1][-1:] in "-_~"):
                    result_parts[-1] += merged
                else:
                    result_parts.append(merged)
                run = []
            if tok:
                result_parts.append(tok)
    if run:
        merged = "".join(run)
        if result_parts and (merged[0] in "-_~" or result_parts[-1][-1:] in "-_~"):
            result_parts[-1] += merged
        else:
            result_parts.append(merged)
    return " ".join(result_parts)

def _축척_텍스트_정리(txt: str) -> str:
    if not txt: return "X"
    u = txt.upper()
    if "NONE" in u or "N/A" in u: return "NONE"
    m = _축척_패턴.search(u)
    return f"1/{m.group(2).replace(',', '')}" if m and m.group(2) else "X"

def _extract_drawing_number(text: str) -> Optional[str]:
    for m in _도면번호_패턴.finditer(text):
        prefix = m.group(1)
        if m.group(0) in ["A1", "A3", "A0", "A2", "A4"]: continue
        exclude_words = ["상세", "일람", "배치", "전개", "마감", "계획", "조감", "구조", "코어", "지하", "옥상", "옥탑", "지붕", "주동", "단위", "세대", "내역", "관계", "형별", "부분", "창호", "가구", "조경", "토목", "기계", "전기", "범례", "개요", "표지", "도면", "시설", "센터", "주차장", "휴게소", "사무소", "경로당", "어린이집", "유치원", "도서관", "커뮤니티", "피트니스", "사우나", "골프", "문주", "경비실"]
        if any(k in prefix for k in exclude_words): continue
        if prefix.endswith("도") or prefix.endswith("표") or prefix.endswith("층") or prefix.endswith("동"): continue
        if len(prefix) > 1 and all("가" <= c <= "힣" for c in prefix): continue
        return m.group(0)
    return None

def _정리문자열(txt: str) -> str:
    return re.sub(r"\s+", " ", (txt or "")).strip()

def _cad_로드(path: Path):
    if path.suffix.lower() == ".dxf": return ezdxf.readfile(str(path))
    _oda_환경_설정()
    from ezdxf.addons import odafc
    return odafc.readfile(str(path))

def _get_safe_point(ent) -> Tuple[float, float]:
    p = ent.dxf.insert
    if getattr(ent.dxf, "halign", 0) > 0 or getattr(ent.dxf, "valign", 0) > 0:
        ap = getattr(ent.dxf, "align_point", None)
        if ap and (round(ap[0], 2) != 0 or round(ap[1], 2) != 0): p = ap
    return float(p[0]), float(p[1])

def _텍스트_데이터_추출(ent) -> List[Tuple[float, float, str, float]]:
    유형 = ent.dxftype(); 결과 = []
    try:
        if 유형 in ["TEXT", "ATTRIB"]:
            px, py = _get_safe_point(ent)
            txt = (ent.dxf.text or "").strip()
            if txt: 결과.append((px, py, txt, float(getattr(ent.dxf, "height", 10.0))))
        elif 유형 == "MTEXT":
            h = getattr(ent.dxf, "char_height", 10.0)
            bx, by = float(ent.dxf.insert[0]), float(ent.dxf.insert[1])
            for i, line in enumerate(ent.plain_text().split('\n')):
                txt = line.strip()
                if txt: 결과.append((bx, by - (i * h * 1.5), txt, float(h)))
        elif 유형 == "ATTDEF":
            px, py = _get_safe_point(ent)
            txt = getattr(ent.dxf, 'tag', '').strip() 
            if not txt: txt = getattr(ent.dxf, 'text', '').strip() 
            if txt: 결과.append((px, py, txt, float(getattr(ent.dxf, "height", 10.0))))
    except Exception as e: logger.debug("텍스트 엔티티 처리 건너뜀: %s", e)
    return 결과

def _collect_layout_texts(layout) -> List[Tuple[float, float, str, float]]:
    texts = []
    try:
        for ent in layout.query("TEXT MTEXT LINE LWPOLYLINE INSERT ATTDEF"):
            if ent.dxftype() in ["TEXT", "MTEXT", "LINE", "LWPOLYLINE", "ATTDEF"]:
                texts.extend(_텍스트_데이터_추출(ent))
            elif ent.dxftype() == "INSERT":
                for att in getattr(ent, "attribs", []): texts.extend(_텍스트_데이터_추출(att))
                try:
                    for v_ent in ent.virtual_entities():
                        if v_ent.dxftype() in ["TEXT", "MTEXT", "LINE", "LWPOLYLINE", "ATTDEF"]: texts.extend(_텍스트_데이터_추출(v_ent))
                        elif v_ent.dxftype() == "INSERT":
                            for v_att in getattr(v_ent, "attribs", []): texts.extend(_텍스트_데이터_추출(v_att))
                except Exception as e: logger.debug("가상 엔티티 처리 건너뜀: %s", e)
    except Exception as e: logger.debug("레이아웃 텍스트 수집 건너뜀: %s", e)
    seen, out = set(), []
    for x, y, txt, h in texts:
        clean = _정리문자열(txt); key = (round(x, 2), round(y, 2), clean)
        if key not in seen: seen.add(key); out.append((float(x), float(y), clean, float(h)))
    return out

def _parse_xref_original(xref_path: str) -> List[Tuple[float, float, str, float]]:
    logger.info("[XREF] 도곽 원본 스캔 중... (%s)", os.path.basename(xref_path))
    try:
        doc = _cad_로드(Path(xref_path)); texts = []
        for ent in doc.modelspace().query("TEXT MTEXT INSERT ATTDEF"):
            if ent.dxftype() in ["TEXT", "MTEXT", "ATTDEF"]: texts.extend(_텍스트_데이터_추출(ent))
            elif ent.dxftype() == "INSERT":
                for att in getattr(ent, "attribs", []): texts.extend(_텍스트_데이터_추출(att))
                try:
                    for v_ent in ent.virtual_entities():
                        if v_ent.dxftype() in ["TEXT", "MTEXT", "ATTDEF"]: texts.extend(_텍스트_데이터_추출(v_ent))
                except Exception as e: logger.debug("XREF 가상 엔티티 처리 건너뜀: %s", e)
        seen, out = set(), []
        for x, y, txt, h in texts:
            clean = _정리문자열(txt); key = (round(x, 2), round(y, 2), clean)
            if key not in seen: seen.add(key); out.append((float(x), float(y), clean, float(h)))
        logger.info("  -> 엑스레이 스캔 성공! %d개의 고정 텍스트 암기 완료.", len(out))
        return out
    except Exception as e: logger.error("XREF 스캔 실패: %s", e); return []

def _transform_xref_texts(xref_texts: List[Tuple[float, float, str, float]], ix: float, iy: float, xscale: float, yscale: float, rot_deg: float) -> List[Tuple[float, float, str, float]]:
    transformed = []; rad = math.radians(rot_deg); cos_val = math.cos(rad); sin_val = math.sin(rad)
    for x, y, txt, h in xref_texts:
        sx = x * xscale; sy = y * yscale
        rx = sx * cos_val - sy * sin_val; ry = sx * sin_val + sy * cos_val
        transformed.append((ix + rx, iy + ry, txt, h * yscale))
    return transformed

def _clean_title_only(title: str) -> str:
    clean = re.sub(r"NONE|N/A|1\s?[/:,]\s?[\d,]+", " ", title, flags=re.I)
    clean = re.sub(r"(?:축척|SCALE)?\s*\(\s*1\s*[:/]\s*\)", " ", clean, flags=re.I)
    clean = re.sub(r"(?:축척|SCALE)\s*1\s*[:/]", " ", clean, flags=re.I)
    return _clean_text_from_headers(clean)

def _extract_scale_smart(cell_texts: List[Tuple[float, float, str, float]], header_a1_x: Optional[float] = None, header_a3_x: Optional[float] = None, is_list_table: bool = False) -> Tuple[str, str]:
    texts_to_scan = []; lone_numbers = []
    if not is_list_table:
        merged_texts = []
        if cell_texts:
            cell_texts_sorted = sorted(cell_texts, key=lambda t: (-t[1], t[0]))
            curr_line = []; curr_y = None
            for t in cell_texts_sorted:
                x, y, txt, h = t
                if curr_y is None: curr_y = y; curr_line.append(t)
                elif abs(curr_y - y) <= max(h * 1.5, 1.0): curr_line.append(t)
                else:
                    curr_line.sort(key=lambda item: item[0])
                    merged_texts.append((curr_line[0][0], curr_y, " ".join([item[2] for item in curr_line]), curr_line[0][3]))
                    curr_y = y; curr_line = [t]
            if curr_line:
                curr_line.sort(key=lambda item: item[0])
                merged_texts.append((curr_line[0][0], curr_y, " ".join([item[2] for item in curr_line]), curr_line[0][3]))
        texts_to_scan = merged_texts
    else: texts_to_scan = cell_texts

    a1_val, a3_val = "X", "X"; scales, labels = [], {}
    for x, y, txt, h in texts_to_scan:
        u_txt = txt.upper(); clean_txt = u_txt.replace(" ", "")
        m_a1 = re.search(r'A1.*?(1\s?[/:,]\s?[\d,]+|NONE|N/A)', clean_txt)
        if m_a1 and a1_val == "X": a1_val = _축척_텍스트_정리(m_a1.group(1))
        m_a3 = re.search(r'A3.*?(1\s?[/:,]\s?[\d,]+|NONE|N/A)', clean_txt)
        if m_a3 and a3_val == "X": a3_val = _축척_텍스트_정리(m_a3.group(1))
        
        if re.search(r'\bA1\b', u_txt): labels['A1'] = (x, y)
        if re.search(r'\bA3\b', u_txt): labels['A3'] = (x, y)
        
        for m in _축척_패턴.finditer(u_txt):
            val = _축척_텍스트_정리(m.group(0))
            if val != "X": scales.append((x, y, val))
            
        if not is_list_table and not _축척_패턴.search(u_txt):
            if re.search(r'^[\d,]+$', clean_txt): lone_numbers.append((x, y, f"1/{clean_txt.replace(',', '')}"))
            elif clean_txt in ["NONE", "N/A"]: lone_numbers.append((x, y, "NONE"))
            
    unique_scales, seen = [], set()
    for sx, sy, sval in scales + lone_numbers:
        if (sx, sy, sval) not in seen: seen.add((sx, sy, sval)); unique_scales.append((sx, sy, sval))
            
    def dist(x1, y1, x2, y2): return math.sqrt((x1 - x2)**2 + (y1 - y2)**2)
    pairings = []
    for sx, sy, sval in unique_scales:
        d_a1 = dist(sx, sy, labels['A1'][0], labels['A1'][1]) if 'A1' in labels else (abs(sx - header_a1_x) if header_a1_x is not None else float('inf'))
        d_a3 = dist(sx, sy, labels['A3'][0], labels['A3'][1]) if 'A3' in labels else (abs(sx - header_a3_x) if header_a3_x is not None else float('inf'))
        if d_a1 != float('inf') or d_a3 != float('inf'):
            if d_a1 <= d_a3: pairings.append((d_a1, sval, 'A1'))
            else: pairings.append((d_a3, sval, 'A3'))
            
    pairings.sort(key=lambda p: p[0])
    for d, sval, target in pairings:
        if target == 'A1' and a1_val == "X": a1_val = sval
        elif target == 'A3' and a3_val == "X": a3_val = sval
        
    if unique_scales:
        unique_scales.sort(key=lambda item: item[0])
        if a1_val == "X" and a3_val == "X":
            if len(unique_scales) >= 2: a1_val, a3_val = unique_scales[0][2], unique_scales[1][2]
            else: a1_val = unique_scales[0][2]
        elif a1_val == "X" and a3_val != "X":
            for _, _, sval in unique_scales:
                if sval != a3_val: a1_val = sval; break
        elif a3_val == "X" and a1_val != "X":
            for _, _, sval in unique_scales:
                if sval != a1_val: a3_val = sval; break
    return a1_val, a3_val

# ============================================================================
# 2. 도면목록표 및 개별 도면 파싱 코어 (Master/Slave 분리 적용)
# ============================================================================
def extract_dwg_list_table(dwg_path: str, block_name: str, roi_cfg: dict, base_w: float, base_h: float, xref_texts: List[Tuple[float, float, str, float]]) -> pd.DataFrame:
    logger.info("[LIST] DWG 도면목록표 분석 시작: %s", os.path.basename(dwg_path))
    데이터, 목표블록 = [], block_name.strip().lower()
    list_rois = roi_cfg.get('list_rois', [])
    global_ignores_stripped = [h.replace(" ", "").upper() for h in GLOBAL_IGNORE_HEADERS]
    
    try:
        doc = _cad_로드(Path(dwg_path))
        for layout in doc.layouts:
            도곽들 = [ins for ins in layout.query("INSERT") if 목표블록 in ins.dxf.name.lower()]
            if not 도곽들: continue
            레이아웃_원본텍스트 = _collect_layout_texts(layout)
            for 도곽 in 도곽들:
                ix, iy = float(도곽.dxf.insert.x), float(도곽.dxf.insert.y)
                xscale, yscale = abs(float(도곽.dxf.xscale)), abs(float(도곽.dxf.yscale))
                너비, 높이 = base_w * xscale, base_h * yscale
                rot_deg = getattr(도곽.dxf, 'rotation', 0.0); rad = math.radians(-rot_deg); cos_val, sin_val = math.cos(rad), math.sin(rad)
                모든텍스트 = 레이아웃_원본텍스트.copy()
                if xref_texts: 모든텍스트.extend(_transform_xref_texts(xref_texts, ix, iy, xscale, yscale, rot_deg))

                target_ranges = list_rois if list_rois else [[0.0, 1.0, 0.0, 1.0]]
                for roi_idx, roi in enumerate(target_ranges):
                    min_x, max_x = ix + (너비 * roi[0]), ix + (너비 * roi[1])
                    y_min, y_max = iy + (높이 * roi[2]), iy + (높이 * roi[3]); roi_w = max_x - min_x
                    num_x_cands, title_x_cands, remark_x_cands, a1_matches, a3_matches, 구역_텍스트 = [], [], [], [], [], []
                    for t in 모든텍스트:
                        tx, ty, txt, th = t
                        dx, dy = tx - ix, ty - iy
                        unrot_x = ix + (dx * cos_val - dy * sin_val); unrot_y = iy + (dx * sin_val + dy * cos_val)
                        if min_x <= unrot_x <= max_x and y_min <= unrot_y <= y_max:
                            clean_t = txt.replace(" ", "").replace("\n", "").strip().upper()
                            if clean_t in ["도면번호", "도연번호", "DWG.NO", "DWG.NO.", "DWGNO", "DRAWINGNO", "번호"]: num_x_cands.append(unrot_x)
                            if clean_t in ["도면명", "DRAWINGTITLE", "TITLE", "도면명칭"]: title_x_cands.append(unrot_x)
                            if clean_t in ["비고", "REMARK", "REMARKS"]: remark_x_cands.append(unrot_x)
                            if txt == "-" and th > roi_w * 0.8: continue
                            if not _extract_drawing_number(txt):
                                if re.search(r"\bA1\b", txt.upper()): a1_matches.append((unrot_x, unrot_y, txt, th))
                                if re.search(r"\bA3\b", txt.upper()): a3_matches.append((unrot_x, unrot_y, txt, th))
                            if any(ih == clean_t for ih in global_ignores_stripped): continue
                            구역_텍스트.append((unrot_x, unrot_y, txt, th))
                    
                    if not 구역_텍스트: continue
                    header_num_x = sum(num_x_cands)/len(num_x_cands) if num_x_cands else min_x + (roi_w * 0.15)
                    header_title_x = sum(title_x_cands)/len(title_x_cands) if title_x_cands else min_x + (roi_w * 0.5)
                    header_remark_x = sum(remark_x_cands)/len(remark_x_cands) if remark_x_cands else max_x
                    header_a1_cands = [m for m in a1_matches if abs(m[0] - header_num_x) > abs(m[0] - header_title_x)]
                    header_a3_cands = [m for m in a3_matches if abs(m[0] - header_num_x) > abs(m[0] - header_title_x)]
                    header_a1_item = sorted(header_a1_cands, key=lambda v: -v[1])[0] if header_a1_cands else None
                    header_a3_item = sorted(header_a3_cands, key=lambda v: -v[1])[0] if header_a3_cands else None
                    if header_a1_item and header_a1_item in 구역_텍스트: 구역_텍스트.remove(header_a1_item)
                    if header_a3_item and header_a3_item in 구역_텍스트: 구역_텍스트.remove(header_a3_item)
                    header_a1_x = header_a1_item[0] if header_a1_item else None
                    header_a3_x = header_a3_item[0] if header_a3_item else None

                    for i in range(len(구역_텍스트)):
                        tx, ty, txt, th = 구역_텍스트[i]
                        if txt.strip() in ["-", "_", "~"]:
                            closest_y = ty; min_dist = float('inf')
                            for j in range(len(구역_텍스트)):
                                if i == j: continue
                                ox, oy, otxt, oth = 구역_텍스트[j]
                                if otxt.strip() not in ["-", "_", "~"]:
                                    if abs(ty - oy) < 높이 * 0.025:
                                        dist_x = abs(tx - ox)
                                        if dist_x < min_dist: min_dist = dist_x; closest_y = oy
                            구역_텍스트[i] = (tx, closest_y, txt, th)

                    구역_텍스트.sort(key=lambda x: -x[1]) 
                    sub_lines, curr_sub, curr_y = [], [], None
                    for t in 구역_텍스트:
                        if curr_y is None or abs(curr_y - t[1]) <= 높이 * 0.012: curr_y = t[1]; curr_sub.append(t)
                        else:
                            curr_sub.sort(key=lambda x: x[0]); sub_lines.append({'y': curr_y, 'texts': curr_sub}); curr_y = t[1]; curr_sub = [t]
                    if curr_sub: curr_sub.sort(key=lambda x: x[0]); sub_lines.append({'y': curr_y, 'texts': curr_sub})

                    rows, unassigned_sub_lines = [], []
                    for sub in sub_lines:
                        full_str = " ".join([t[2] for t in sub['texts']])
                        is_category = False
                        if any(kw in full_str.replace(" ", "") for kw in CATEGORY_KEYWORDS): is_category = True
                        elif re.search(r"^[A-Z0-9\-_]*\s*[\[<【].+?[\]>】]\s*$", full_str): is_category = True
                        if is_category: continue

                        num_texts = [t for t in sub['texts'] if abs(t[0] - header_num_x) <= abs(t[0] - header_title_x)]
                        num_str = _spatial_reconstruct_num_str(num_texts)
                        raw_drw_no = _extract_drawing_number(num_str) or _extract_drawing_number(full_str)
                        drw_no, raw_matched_str = "", ""
                        if raw_drw_no: drw_no = _도면번호_세척(raw_drw_no); raw_matched_str = raw_drw_no
                        else:
                            if num_texts:
                                fallback_match = re.sub(r"\s*[가-힣\[<【\(].*$", "", num_str).strip("-_ ")
                                if not fallback_match: fallback_match = num_str.strip()
                                if re.search(r"\d", fallback_match) and len(fallback_match) >= 3 and not re.search(r"[\[\]<>\(【】]", fallback_match):
                                    drw_no = _도면번호_세척(fallback_match); raw_matched_str = fallback_match

                        if drw_no: rows.append({'anchor_y': sub['y'], 'sub_lines': [{'y': sub['y'], 'texts': sub['texts'], 'raw_drw_no': raw_matched_str}], 'drw_no': drw_no})
                        else: unassigned_sub_lines.append({'y': sub['y'], 'texts': sub['texts']})

                    for sub in unassigned_sub_lines:
                        if not rows: continue
                        closest_row = min(rows, key=lambda r: abs(r['anchor_y'] - sub['y']))
                        if abs(closest_row['anchor_y'] - sub['y']) < 높이 * 0.04: closest_row['sub_lines'].append(sub)

                    avg_char_h = (sum(t[3] for t in 구역_텍스트) / len(구역_텍스트)) if 구역_텍스트 else 1.0
                    prop_dong = "공통"; dong_title_ref_x = None
                    for row in rows:
                        row['sub_lines'].sort(key=lambda s: -s['y']); title_words, all_texts = [], []; row_min_title_x = None
                        for sub in row['sub_lines']:
                            sub_texts_sorted = sorted(sub['texts'], key=lambda x: x[0]); title_texts = []
                            for t in sub_texts_sorted:
                                if header_remark_x and abs(t[0] - header_remark_x) < abs(t[0] - header_title_x): continue
                                title_texts.append(t)
                            if title_texts:
                                sub_min_x = min(t[0] for t in title_texts)
                                if row_min_title_x is None or sub_min_x < row_min_title_x: row_min_title_x = sub_min_x
                            raw_left_str = _spatial_reconstruct_num_str(title_texts)
                            title_overflow = raw_left_str
                            if sub.get('raw_drw_no') and sub['raw_drw_no'] in raw_left_str:
                                parts = raw_left_str.split(sub['raw_drw_no'], 1); title_overflow = parts[1] if len(parts) > 1 else ""
                            title_overflow = _merge_title_char_runs(title_overflow)
                            cleaned_line = _clean_title_only(title_overflow)
                            if cleaned_line: title_words.append(cleaned_line)
                            all_texts.extend(sub_texts_sorted)

                        번호 = row['drw_no']; 명칭 = " ".join(title_words).strip()
                        current_dong = "공통"; extracted_dong = _extract_dong_from_title(명칭)
                        if extracted_dong and 명칭.lstrip().startswith(extracted_dong):
                            current_dong = extracted_dong; 임시_명칭 = 명칭.replace(current_dong, "")
                            임시_명칭 = re.sub(r"^[,\s]+|[,\s]+$", "", 임시_명칭).strip()
                            if 임시_명칭: 명칭 = 임시_명칭
                            gp = _extract_group_prefix(명칭)
                            if gp:
                                current_dong = current_dong + " " + gp
                                명칭 = re.sub(r"^" + re.escape(gp) + r"[,\s]*", "", 명칭).strip()
                            prop_dong = current_dong; dong_title_ref_x = row_min_title_x
                        elif (dong_title_ref_x is not None and row_min_title_x is not None
                              and row_min_title_x > dong_title_ref_x + avg_char_h * 1.5):
                            current_dong = prop_dong
                        else:
                            gp = _extract_group_prefix(명칭)
                            if gp: prop_dong = gp; dong_title_ref_x = row_min_title_x
                            else: prop_dong = "공통"; dong_title_ref_x = None

                        a1, a3 = _extract_scale_smart(all_texts, header_a1_x, header_a3_x, is_list_table=True)
                        데이터.append({"도면번호(LIST)": 번호, "구분_LIST(동)": current_dong if current_dong != "공통" else "", "도면명(LIST)": 명칭, "축척_A1(LIST)": a1, "축척_A3(LIST)": a3})
    except Exception as e: logger.error("목록표 분석 중 오류: %s", e)
    df = pd.DataFrame(데이터)
    return pd.DataFrame(columns=["도면번호(LIST)", "구분_LIST(동)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)"]) if df.empty else df.drop_duplicates(subset=["도면번호(LIST)"]).reset_index(drop=True)

def _process_single_dwg(args: Tuple[str, str, dict, float, float, List[Tuple[float, float, str, float]]]) -> Tuple[List[dict], str]:
    전체경로, 목표블록, roi_cfg, base_w, base_h, xref_texts = args
    파일명, 데이터, 에러메시지 = os.path.basename(전체경로), [], ""
    try:
        doc = _cad_로드(Path(전체경로)); 도곽_발견됨 = False
        for layout in doc.layouts:
            # [핵심] 사용자가 지정한 개별도면용 '목표블록' 이름(Slave)으로 찾습니다.
            도곽들 = [ins for ins in layout.query("INSERT") if 목표블록 in ins.dxf.name.lower()]
            if not 도곽들: continue
            도곽_발견됨 = True; 레이아웃_원본텍스트 = _collect_layout_texts(layout)
            for 도곽 in 도곽들:
                ix, iy = float(도곽.dxf.insert.x), float(도곽.dxf.insert.y)
                xscale, yscale = abs(float(도곽.dxf.xscale)), abs(float(도곽.dxf.yscale))
                너비, 높이 = base_w * xscale, base_h * yscale
                rot_deg = getattr(도곽.dxf, 'rotation', 0.0); rad = math.radians(-rot_deg); cos_val, sin_val = math.cos(rad), math.sin(rad)
                모든텍스트 = 레이아웃_원본텍스트.copy()
                if xref_texts: 모든텍스트.extend(_transform_xref_texts(xref_texts, ix, iy, xscale, yscale, rot_deg))

                def get_data_in_roi(roi):
                    # [핵심] 비율은 Master의 roi_cfg에 저장된 것을 그대로 씁니다.
                    x_min, x_max = ix + (너비 * roi[0]), ix + (너비 * roi[1])
                    y_min, y_max = iy + (높이 * roi[2]), iy + (높이 * roi[3]); 박스내글자 = []
                    for t in 모든텍스트:
                        tx, ty, txt, th = t
                        dx, dy = tx - ix, ty - iy
                        unrot_x = ix + (dx * cos_val - dy * sin_val); unrot_y = iy + (dx * sin_val + dy * cos_val)
                        if x_min <= unrot_x <= x_max and y_min <= unrot_y <= y_max:
                            if txt == "-" and th > (x_max - x_min) * 0.8: continue
                            박스내글자.append((unrot_x, unrot_y, txt, th)) 
                    if not 박스내글자: return "", []
                    for i in range(len(박스내글자)):
                        tx, ty, txt, th = 박스내글자[i]
                        if txt.strip() in ["-", "_", "~"]:
                            closest_y = ty; min_dist = float('inf')
                            for j in range(len(박스내글자)):
                                if i == j: continue
                                ox, oy, otxt, oth = 박스내글자[j]
                                if otxt.strip() not in ["-", "_", "~"]:
                                    if abs(ty - oy) < 높이 * 0.025:
                                        dist_x = abs(tx - ox)
                                        if dist_x < min_dist: min_dist = dist_x; closest_y = oy
                            박스내글자[i] = (tx, closest_y, txt, th)
                    박스내글자.sort(key=lambda t: -t[1])
                    lines, current_line, current_y = [], [], None
                    for t in 박스내글자:
                        if current_y is None: current_y = t[1]; current_line.append(t)
                        elif abs(current_y - t[1]) <= 높이 * 0.015: current_line.append(t)
                        else:
                            current_line.sort(key=lambda x: x[0]); lines.append(" ".join([x[2] for x in current_line])); current_y = t[1]; current_line = [t]
                    if current_line: current_line.sort(key=lambda x: x[0]); lines.append(" ".join([x[2] for x in current_line]))
                    return " ".join(lines), 박스내글자

                n_str, _ = get_data_in_roi(roi_cfg['num_roi']); t_str, _ = get_data_in_roi(roi_cfg['title_roi']); _, s_texts = get_data_in_roi(roi_cfg['scale_roi']) 
                n_str_clean = _clean_text_from_headers(n_str); t_str_clean = _clean_text_from_headers(t_str)

                번호_후보 = _extract_drawing_number(n_str_clean); raw_matched_str = ""
                if 번호_후보: 번호 = _도면번호_세척(번호_후보); raw_matched_str = 번호_후보
                else: 
                    fallback_match = re.sub(r"\s*[가-힣\[<【\(].*$", "", n_str_clean).strip("-_ ")
                    번호 = _도면번호_세척(fallback_match); raw_matched_str = fallback_match
                
                명칭 = t_str_clean
                if raw_matched_str and raw_matched_str in 명칭: 명칭 = 명칭.replace(raw_matched_str, "")
                dwg_dong = _extract_dong_from_title(명칭)
                if dwg_dong:
                    임시_명칭 = 명칭.replace(dwg_dong, "")
                    임시_명칭 = re.sub(r"^[,\s]+|[,\s]+$", "", 임시_명칭).strip()
                    if 임시_명칭: 명칭 = 임시_명칭
                    
                명칭 = _clean_title_only(명칭); a1, a3 = _extract_scale_smart(s_texts, is_list_table=False)
                if 번호: 데이터.append({"파일명": 파일명, "도면번호(DWG)": 번호, "구분_DWG(동)": dwg_dong, "도면명(DWG)": 명칭.strip(), "축척_A1(DWG)": a1, "축척_A3(DWG)": a3})
        del doc
        if not 도곽_발견됨: return 데이터, "도곽 블록 없음"
    except Exception as e: 에러메시지 = str(e)
    return 데이터, 에러메시지

def extract_dwg_data_multiprocess(target_dirs: List[str], slave_block_name: str, roi_cfg: dict, base_w: float, base_h: float, xref_texts: List[Tuple[float, float, str, float]]) -> pd.DataFrame:
    모든_캐드파일 = []
    for d in target_dirs:
        폴더 = Path(d)
        if 폴더.exists(): 모든_캐드파일.extend([str(p) for p in 폴더.iterdir() if p.is_file() and p.suffix.lower() in [".dwg", ".dxf"]])
    캐드파일들 = sorted(list(set(모든_캐드파일)))
    if not 캐드파일들:
        logger.warning("[CAD ] 폴더 내에 처리할 도면 파일이 없습니다."); return pd.DataFrame(columns=["파일명", "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])

    logger.info("[CAD ] 총 %d개의 개별 도면 분석 중... (터보 모드 가동 🚀)", len(캐드파일들))
    최종_데이터 = []
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {executor.submit(_process_single_dwg, (path, slave_block_name.strip().lower(), roi_cfg, base_w, base_h, xref_texts)): path for path in 캐드파일들}
        for i, future in enumerate(concurrent.futures.as_completed(futures), 1):
            경로 = futures[future]
            try:
                결과, 에러 = future.result()
                if 결과: 최종_데이터.extend(결과)
                logger.info("   [%d/%d] %s: %s (%s)", i, len(캐드파일들), '완료' if 결과 else '패스', os.path.basename(경로), 에러 if 에러 else '성공')
            except Exception as e: logger.error("   [%d/%d] 시스템 오류: %s (%s)", i, len(캐드파일들), os.path.basename(경로), e)
    
    if not 최종_데이터: return pd.DataFrame(columns=["파일명", "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])
    return pd.DataFrame(최종_데이터)

def build_report(list_df: pd.DataFrame, dwg_df: pd.DataFrame, out_path: str):
    if list_df.empty and dwg_df.empty: logger.warning("[알림] 추출된 데이터가 없어 엑셀 리포트를 생성하지 않습니다."); return

    lst, dwg = list_df.copy(), dwg_df.copy()
    if "도면번호(LIST)" not in lst.columns: lst["도면번호(LIST)"] = ""
    if "도면번호(DWG)" not in dwg.columns: dwg["도면번호(DWG)"] = ""
    if "구분_LIST(동)" not in lst.columns: lst["구분_LIST(동)"] = ""
    if "구분_DWG(동)" not in dwg.columns: dwg["구분_DWG(동)"] = ""

    lst["KEY"] = lst["도면번호(LIST)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    dwg["KEY"] = dwg["도면번호(DWG)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    결과 = pd.merge(lst, dwg, on="KEY", how="outer", indicator=True)
    결과["상태"] = 결과["_merge"].map({"both": "일치", "left_only": "DWG 누락", "right_only": "목록표 누락"})

    dong_mismatch_indices = set()
    for i in range(len(결과)):
        l_d = str(결과.at[i, "구분_LIST(동)"]).strip(); d_d = str(결과.at[i, "구분_DWG(동)"]).strip()
        if l_d == "nan": l_d = ""
        if d_d == "nan": d_d = ""
        if l_d and d_d and l_d != d_d: dong_mismatch_indices.add(i + 2)

    prev_dong = ""; dong_col_idx = 결과.columns.get_loc("구분_LIST(동)")
    for i in range(len(결과)):
        curr_dong = str(결과.iat[i, dong_col_idx]).strip()
        if curr_dong == "nan" or not curr_dong: prev_dong = ""; 결과.iat[i, dong_col_idx] = ""; continue
        if curr_dong == prev_dong: 결과.iat[i, dong_col_idx] = ""  
        else: prev_dong = curr_dong          

    prev_dwg_dong = ""; dwg_dong_col_idx = 결과.columns.get_loc("구분_DWG(동)")
    for i in range(len(결과)):
        curr_dong = str(결과.iat[i, dwg_dong_col_idx]).strip()
        if curr_dong == "nan" or not curr_dong: prev_dwg_dong = ""; 결과.iat[i, dwg_dong_col_idx] = ""; continue
        if curr_dong == prev_dwg_dong: 결과.iat[i, dwg_dong_col_idx] = ""  
        else: prev_dwg_dong = curr_dong          

    cols = ["도면번호(LIST)", "구분_LIST(동)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)", 
            "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)", "파일명", "상태"]
    for c in cols: 
        if c not in 결과.columns: 결과[c] = ""
    
    결과[cols].fillna("X").to_excel(out_path, index=False)
    wb = load_workbook(out_path); ws = wb.active
    빨간색 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    h = {cell.value: cell.column for cell in ws[1] if cell.value}
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, h["상태"]).value != "일치":
            for c in range(1, len(cols)+1): ws.cell(row, c).fill = 빨간색
        else:
            if row in dong_mismatch_indices:
                if h.get("구분_LIST(동)"): ws.cell(row, h.get("구분_LIST(동)")).fill = 빨간색
                if h.get("구분_DWG(동)"): ws.cell(row, h.get("구분_DWG(동)")).fill = 빨간색
            val_list = re.sub(r"[\s\-_]", "", str(ws.cell(row, h.get("도면번호(LIST)")).value).upper())
            val_dwg = re.sub(r"[\s\-_]", "", str(ws.cell(row, h.get("도면번호(DWG)")).value).upper())
            if val_list != val_dwg:
                ws.cell(row, h.get("도면번호(LIST)")).fill = 빨간색
                ws.cell(row, h.get("도면번호(DWG)")).fill = 빨간색
            name_list = str(ws.cell(row, h.get("도면명(LIST)")).value).replace(" ", "")
            name_dwg = str(ws.cell(row, h.get("도면명(DWG)")).value).replace(" ", "")
            if name_list != name_dwg:
                ws.cell(row, h.get("도면명(LIST)")).fill = 빨간색
                ws.cell(row, h.get("도면명(DWG)")).fill = 빨간색
            for s in ["A1", "A3"]:
                p_v = str(ws.cell(row, h[f"축척_{s}(LIST)"]).value).replace(" ","")
                d_v = str(ws.cell(row, h[f"축척_{s}(DWG)"]).value).replace(" ","")
                if p_v != d_v:
                    ws.cell(row, h[f"축척_{s}(LIST)"]).fill = 빨간색
                    ws.cell(row, h[f"축척_{s}(DWG)"]).fill = 빨간색

    wb.save(out_path)
    logger.info("[XLSX] 리포트 저장 완료: %s", out_path)

# ============================================================================
# 3. [GUI 구축] CustomTkinter + TkinterDnD (드래그 앤 드롭 지원)
# ============================================================================
class GUILogHandler(logging.Handler):
    """로그 레코드를 GUI 텍스트박스에 출력하는 핸들러."""
    def __init__(self, textbox: ctk.CTkTextbox):
        super().__init__()
        self.textbox = textbox

    def emit(self, record: logging.LogRecord):
        msg = self.format(record)
        self.textbox.configure(state="normal")
        self.textbox.insert("end", msg + "\n")
        self.textbox.see("end")
        self.textbox.configure(state="disabled")

# [핵심] ctk.CTk와 TkinterDnD.DnDWrapper를 결합하여 D&D 윈도우 생성
class AutoDWGApp(ctk.CTk, TkinterDnD.DnDWrapper):
    def __init__(self):
        super().__init__()
        self.TkdndVersion = TkinterDnD._require(self)
        
        self.title("AutoDWG 도면 검토 자동화 시스템")
        self.geometry("760x860")
        self.resizable(False, False)
        
        self.xref_path = ""
        self.list_path = ""
        self.dwg_folders = []

        self._build_ui()

        gui_handler = GUILogHandler(self.log_box)
        gui_handler.setLevel(logging.DEBUG)
        gui_handler.setFormatter(logging.Formatter("%(message)s"))
        logger.addHandler(gui_handler)

        logger.info("=" * 72)
        logger.info(" AutoDWG 검토 자동화 시스템  v6.7 _ © 2026. 김정현. All rights reserved.")
        logger.info("=" * 72)
        logger.info(" 환영합니다! 파일이나 폴더를 '드래그 앤 드롭' 하거나 버튼으로 추가하세요.\n")

    def _parse_dnd_paths(self, dnd_data):
        # 윈도우 탐색기에서 넘어온 복잡한 경로 문자열({} 포함 등)을 깔끔한 리스트로 분리
        return self.tk.splitlist(dnd_data)

    def _build_ui(self):
        F_H1   = ctk.CTkFont(family="Malgun Gothic", size=20, weight="bold")
        F_H2   = ctk.CTkFont(family="Malgun Gothic", size=12, weight="bold")
        F_BODY = ctk.CTkFont(family="Malgun Gothic", size=12)
        F_SM   = ctk.CTkFont(family="Malgun Gothic", size=10)
        F_BTN  = ctk.CTkFont(family="Malgun Gothic", size=14, weight="bold")
        F_MONO = ctk.CTkFont(family="Consolas", size=11)

        BG      = "#f2f4f8"
        CARD    = "#ffffff"
        HDRBG   = "#e8f0fd"
        PRIMARY = "#1e5fbe"
        DANGER  = "#d94f43"
        DIM     = "#8a94a6"

        self.configure(fg_color=BG)

        topbar = ctk.CTkFrame(self, fg_color=PRIMARY, corner_radius=0, height=58)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        inn = ctk.CTkFrame(topbar, fg_color="transparent")
        inn.pack(fill="both", expand=True, padx=22)
        ctk.CTkLabel(inn, text="AutoDWG  검토 자동화 시스템",
                     font=F_H1, text_color="white", anchor="w").pack(side="left", fill="y")
        ctk.CTkLabel(inn, text="v 6.7 _ © 2026. 김정현. All rights reserved.", font=F_SM,
                     text_color="#a8c8ff", anchor="e").pack(side="right")

        page = ctk.CTkScrollableFrame(self, fg_color=BG,
                                      scrollbar_button_color="#c0c8d8",
                                      scrollbar_button_hover_color="#a8b4cc")
        page.pack(fill="both", expand=True)

        def _card(heading):
            outer = ctk.CTkFrame(page, fg_color=CARD, corner_radius=10)
            outer.pack(fill="x", padx=16, pady=5)
            hdr = ctk.CTkFrame(outer, fg_color=HDRBG, corner_radius=8, height=30)
            hdr.pack(fill="x", padx=1, pady=(1, 0))
            hdr.pack_propagate(False)
            ctk.CTkLabel(hdr, text=heading, font=F_H2,
                         text_color=PRIMARY, anchor="w").pack(side="left", padx=12)
            body = ctk.CTkFrame(outer, fg_color="transparent")
            body.pack(fill="x", padx=12, pady=(6, 10))
            return outer, body

        def _file_row(parent, label, cmd):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(row, text=label, font=F_H2,
                         width=120, anchor="w").pack(side="left")
            ctk.CTkButton(row, text="파일 선택", font=F_BODY,
                          width=88, height=28, command=cmd).pack(side="left", padx=(0, 10))
            lbl = ctk.CTkLabel(row, text="선택된 파일 없음",
                               font=F_BODY, text_color=DIM, anchor="w")
            lbl.pack(side="left", fill="x", expand=True)
            return lbl

        c1, b1 = _card("①  도곽 원본 (XREF)  및  블록 이름")
        c1.drop_target_register(DND_FILES)
        c1.dnd_bind("<<Drop>>", self.drop_xref)
        self.lbl_xref = _file_row(b1, "원본 DWG", self.select_xref)
        row_blk = ctk.CTkFrame(b1, fg_color="transparent")
        row_blk.pack(fill="x", pady=3)
        ctk.CTkLabel(row_blk, text="블록 이름", font=F_H2,
                     width=120, anchor="w").pack(side="left")
        self.entry_block_name = ctk.CTkEntry(row_blk, font=F_BODY, width=300, height=28,
                                              placeholder_text="파일 선택 시 자동 입력됨")
        self.entry_block_name.pack(side="left")

        c2, b2 = _card("②  도면목록표  DWG")
        c2.drop_target_register(DND_FILES)
        c2.dnd_bind("<<Drop>>", self.drop_list)
        self.lbl_list = _file_row(b2, "목록표 DWG", self.select_list)

        c3, b3 = _card("③  개별 도면 폴더")
        btn_r = ctk.CTkFrame(b3, fg_color="transparent")
        btn_r.pack(fill="x", pady=(0, 5))
        ctk.CTkButton(btn_r, text="＋  폴더 추가", font=F_BODY, width=110, height=28,
                      command=self.add_folder).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_r, text="초기화", font=F_BODY, width=68, height=28,
                      fg_color=DANGER, hover_color="#b5362c",
                      command=self.clear_folders).pack(side="left")
        self.textbox_folders = ctk.CTkTextbox(b3, height=72, font=F_SM,
                                               fg_color="#f7f8fa", corner_radius=6,
                                               border_width=1, border_color="#dde2ea")
        self.textbox_folders.pack(fill="x")
        self.textbox_folders.insert("1.0", "폴더를 추가하거나 이 영역에 폴더를 끌어다 놓으세요.")
        self.textbox_folders.configure(state="disabled")
        c3.drop_target_register(DND_FILES)
        c3.dnd_bind("<<Drop>>", self.drop_folders)
        self.textbox_folders.drop_target_register(DND_FILES)
        self.textbox_folders.dnd_bind("<<Drop>>", self.drop_folders)

        _, b4 = _card("④  고급 옵션")
        self.check_diff_name = ctk.CTkCheckBox(
            b4, text="개별 도면의 도곽 이름이 도면목록표와 다를 경우 체크",
            font=F_BODY, command=self.toggle_diff_name)
        self.check_diff_name.pack(anchor="w", pady=2)
        self.frame_diff = ctk.CTkFrame(b4, fg_color="transparent")
        ctk.CTkLabel(self.frame_diff, text="↳  개별도면 도곽 이름 :",
                     font=F_H2, text_color=DANGER).pack(side="left", padx=(20, 10))
        self.entry_slave_block = ctk.CTkEntry(self.frame_diff, font=F_BODY, width=230, height=28,
                                               placeholder_text="예: XR-form")
        self.entry_slave_block.pack(side="left")

        wrap = ctk.CTkFrame(page, fg_color="transparent")
        wrap.pack(fill="x", padx=16, pady=(8, 0))
        self.btn_start = ctk.CTkButton(wrap, text="검토 시작  →",
                                        font=F_BTN, height=46, corner_radius=8,
                                        fg_color=PRIMARY, hover_color="#174da8",
                                        command=self.start_process)
        self.btn_start.pack(fill="x")

        self.progressbar = ctk.CTkProgressBar(page, mode="indeterminate", height=5,
                                               fg_color="#dde4ef", progress_color=PRIMARY,
                                               corner_radius=2)

        _, blog = _card("작업 로그")
        self.log_box = ctk.CTkTextbox(blog, height=200, font=F_MONO,
                                       fg_color="#1e1e2e", text_color="#cdd6f4",
                                       corner_radius=6)
        self.log_box.pack(fill="both", expand=True)
        self.log_box.configure(state="disabled")

        ctk.CTkFrame(page, fg_color="transparent", height=10).pack()

    # ================= UI 로직 =================
    def toggle_diff_name(self):
        if self.check_diff_name.get(): self.frame_diff.pack(fill="x", after=self.check_diff_name, pady=5)
        else: self.frame_diff.pack_forget()

    # ================= D&D 핸들러 =================
    def drop_xref(self, event):
        paths = self._parse_dnd_paths(event.data)
        if paths and paths[0].lower().endswith(('.dwg', '.dxf')):
            self.xref_path = paths[0]
            self.lbl_xref.configure(text=os.path.basename(self.xref_path), text_color="black")
            base_name = os.path.splitext(os.path.basename(self.xref_path))[0]
            self.entry_block_name.delete(0, "end"); self.entry_block_name.insert(0, base_name)
            logger.info("[알림] 드래그 앤 드롭: 원본 도곽 이름(%s) 자동 입력됨.", base_name)

    def drop_list(self, event):
        paths = self._parse_dnd_paths(event.data)
        if paths and paths[0].lower().endswith(('.dwg', '.dxf')):
            self.list_path = paths[0]
            self.lbl_list.configure(text=os.path.basename(self.list_path), text_color="blue")
            logger.info("[알림] 드래그 앤 드롭: 도면목록표 인식 완료.")

    def drop_folders(self, event):
        paths = self._parse_dnd_paths(event.data)
        for p in paths:
            if os.path.isdir(p) and p not in self.dwg_folders:
                self.dwg_folders.append(p)
        self.update_folder_textbox()

    # ================= 기존 버튼 핸들러 =================
    def select_xref(self):
        path = filedialog.askopenfilename(title="외부참조(XREF) 파일 선택", filetypes=[("AutoCAD Files", "*.dwg *.dxf")])
        if path:
            self.xref_path = path
            self.lbl_xref.configure(text=os.path.basename(path), text_color="black")
            base_name = os.path.splitext(os.path.basename(path))[0]
            self.entry_block_name.delete(0, "end"); self.entry_block_name.insert(0, base_name)
            logger.info("[알림] 원본 파일명 기반으로 도곽 블록 이름(%s)이 자동 입력되었습니다.", base_name)

    def select_list(self):
        path = filedialog.askopenfilename(title="도면목록표 파일 선택", filetypes=[("AutoCAD Files", "*.dwg *.dxf")])
        if path:
            self.list_path = path
            self.lbl_list.configure(text=os.path.basename(path), text_color="blue")

    def add_folder(self):
        path = filedialog.askdirectory(title="개별 도면이 있는 폴더 선택")
        if path and path not in self.dwg_folders:
            self.dwg_folders.append(path)
            self.update_folder_textbox()

    def clear_folders(self):
        self.dwg_folders.clear()
        self.update_folder_textbox()

    def update_folder_textbox(self):
        self.textbox_folders.configure(state="normal")
        self.textbox_folders.delete("1.0", "end")
        if not self.dwg_folders: self.textbox_folders.insert("1.0", "여기에 폴더를 끌어다 놓으세요.")
        else:
            for i, f in enumerate(self.dwg_folders, 1): self.textbox_folders.insert("end", f"[{i}] {f}\n")
        self.textbox_folders.configure(state="disabled")

    # ================= 실행 로직 =================
    def start_process(self):
        master_blk = self.entry_block_name.get().strip()
        slave_blk = master_blk

        if not master_blk: messagebox.showwarning("입력 오류", "공통 도곽 블록 이름을 입력하세요."); return
        if not self.list_path: messagebox.showwarning("입력 오류", "도면목록표 파일을 선택하세요."); return
        if not self.dwg_folders: messagebox.showwarning("입력 오류", "분석할 개별 도면 폴더를 하나 이상 추가하세요."); return
        
        # [스마트 분기] 체크박스가 켜져 있다면, slave(개별도면) 블록 이름을 별도로 챙김
        if self.check_diff_name.get():
            slave_blk = self.entry_slave_block.get().strip()
            if not slave_blk: messagebox.showwarning("입력 오류", "개별도면용 도곽 이름을 입력하세요."); return

        self.btn_start.configure(state="disabled", text="분석 진행 중...  ⏳")
        self.progressbar.pack(fill="x", padx=16, pady=(0, 2))
        self.progressbar.start()
        thread = threading.Thread(target=self.run_core_logic, args=(master_blk, slave_blk), daemon=True)
        thread.start()

    def run_core_logic(self, master_blk, slave_blk):
        try:
            # 1. 박스 좌표(ROI)는 무조건 Master 기준(도면목록표)으로 불러옵니다.
            roi_config = load_roi_config(master_blk)
            if not roi_config:
                logger.error("[오류] '%s'에 대한 구역 설정(JSON) 파일이 없습니다!", master_blk)
                logger.error("캐드에서 목록표 파일에 SET_ROI 명령어로 구역을 지정해 주세요.")
                return

            base_w = float(roi_config.get('base_w', 841.0))
            base_h = float(roi_config.get('base_h', 594.0))
            logger.info("[성공] '%s' 설정을 로드했습니다. (원본크기: %.0fx%.0f)", master_blk, base_w, base_h)

            xref_texts = []
            if self.xref_path and os.path.isfile(self.xref_path):
                xref_texts = _parse_xref_original(self.xref_path)

            logger.info("-" * 72)
            실행폴더 = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            최종_저장경로 = os.path.join(실행폴더, 리포트_이름)

            # 2. 목록표 스캔 (Master 블록 이름 사용)
            list_데이터 = extract_dwg_list_table(self.list_path, master_blk, roi_config, base_w, base_h, xref_texts)

            # 3. 개별도면 스캔 (Slave 블록 이름 사용 - 같으면 Master 이름)
            if master_blk != slave_blk: logger.info("💡 [스마트 탐색 모드] 개별도면은 '%s' 도곽 이름으로 탐색을 시작합니다.", slave_blk)
            dwg_데이터 = extract_dwg_data_multiprocess(self.dwg_folders, slave_blk, roi_config, base_w, base_h, xref_texts)

            # 4. 리포트 생성
            build_report(list_데이터, dwg_데이터, 최종_저장경로)

            logger.info("-" * 72)
            logger.info("[DONE] 검토 완료! 리포트가 프로그램과 같은 폴더에 저장되었습니다.")
            os.startfile(실행폴더)

        except PermissionError: logger.error("[ERROR] 엑셀 파일이 이미 켜져 있습니다. 창을 닫고 다시 실행해 주세요.")
        except Exception as e: logger.error("[ERROR] 시스템 오류 발생: %s", e)
        finally:
            def _finish():
                self.btn_start.configure(state="normal", text="검토 시작  →")
                self.progressbar.stop()
                self.progressbar.pack_forget()
            self.after(0, _finish)

# ============================================================================
# 메인 실행
# ============================================================================
if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    
    if not _oda_환경_설정():
        root = tk.Tk(); root.withdraw()
        msg = "⚠️ CAD 분석 엔진(ODA)이 설치되어 있지 않습니다!\n\n확인을 누르면 다운로드 페이지가 열립니다."
        if messagebox.askokcancel("엔진 설치 안내", msg): webbrowser.open(ODA_DOWNLOAD_URL)
        sys.exit()

    app = AutoDWGApp()
    app.mainloop()