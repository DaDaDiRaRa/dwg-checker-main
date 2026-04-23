"""
app.py 핵심 함수 단위 테스트
GUI 의존성(customtkinter, tkinterdnd2)은 임포트 전에 mock 처리합니다.
"""
import sys
import math
import tempfile
import os
import unittest
from unittest.mock import MagicMock

# GUI 전용 패키지 mock (디스플레이 없는 headless 환경에서도 실행 가능)
# tkinter: 기본 mock
_tk_mock = MagicMock()
_tk_mock.Tk = type('Tk', (), {})
sys.modules['tkinter'] = _tk_mock
sys.modules['tkinter.filedialog'] = MagicMock()
sys.modules['tkinter.messagebox'] = MagicMock()

# customtkinter: CTk를 실제 클래스 stub으로 제공 (상속 가능하도록)
class _FakeCTk:
    def __init__(self, *a, **kw): pass
_ctk_mock = MagicMock()
_ctk_mock.CTk = _FakeCTk
sys.modules['customtkinter'] = _ctk_mock

# tkinterdnd2: DnDWrapper를 실제 클래스 stub으로 제공
class _FakeDnDWrapper:
    pass
_dnd_tkinter = MagicMock()
_dnd_tkinter.DnDWrapper = _FakeDnDWrapper
_dnd_tkinter._require = MagicMock(return_value="")
_dnd_mock = MagicMock()
_dnd_mock.TkinterDnD = _dnd_tkinter
_dnd_mock.DND_FILES = "DND_FILES"
sys.modules['tkinterdnd2'] = _dnd_mock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import (
    _정리문자열,
    _도면번호_세척,
    _축척_텍스트_정리,
    _extract_drawing_number,
    _extract_dong_from_title,
    _clean_text_from_headers,
    _transform_xref_texts,
    build_report,
)
import pandas as pd


# ============================================================================
# _정리문자열
# ============================================================================
class TestCleanString(unittest.TestCase):
    def test_strips_whitespace(self):
        self.assertEqual(_정리문자열("  hello  "), "hello")

    def test_collapses_internal_spaces(self):
        self.assertEqual(_정리문자열("a  b   c"), "a b c")

    def test_empty_string(self):
        self.assertEqual(_정리문자열(""), "")

    def test_none_treated_as_empty(self):
        self.assertEqual(_정리문자열(None), "")

    def test_newlines_collapsed(self):
        self.assertEqual(_정리문자열("a\nb"), "a b")


# ============================================================================
# _도면번호_세척
# ============================================================================
class TestDrawingNumberClean(unittest.TestCase):
    def test_empty(self):
        self.assertEqual(_도면번호_세척(""), "")

    def test_uppercase(self):
        self.assertEqual(_도면번호_세척("aa-001"), "AA-001")

    def test_toe_replacement(self):
        result = _도면번호_세척("AA-TOE")
        self.assertIn("108", result)

    def test_leading_dot(self):
        result = _도면번호_세척(".123")
        self.assertTrue(result.startswith("AA"))

    def test_spaces_around_hyphen_collapsed(self):
        self.assertEqual(_도면번호_세척("AA - 001"), "AA-001")

    def test_multiple_hyphens_collapsed(self):
        self.assertEqual(_도면번호_세척("AA--001"), "AA-001")

    def test_greek_lambda_to_a(self):
        self.assertEqual(_도면번호_세척("Λ-001"), "A-001")

    def test_greek_delta_to_a(self):
        self.assertEqual(_도면번호_세척("Δ-001"), "A-001")

    def test_trailing_whitespace_stripped(self):
        self.assertEqual(_도면번호_세척("  AA-001  "), "AA-001")

    def test_underscore_treated_as_hyphen(self):
        result = _도면번호_세척("AA__001")
        self.assertNotIn("__", result)


# ============================================================================
# _축척_텍스트_정리
# ============================================================================
class TestScaleTextClean(unittest.TestCase):
    def test_empty_returns_x(self):
        self.assertEqual(_축척_텍스트_정리(""), "X")

    def test_colon_format(self):
        self.assertEqual(_축척_텍스트_정리("1:100"), "1/100")

    def test_slash_format_unchanged(self):
        self.assertEqual(_축척_텍스트_정리("1/200"), "1/200")

    def test_comma_separator(self):
        self.assertEqual(_축척_텍스트_정리("1,200"), "1/200")

    def test_none_keyword(self):
        self.assertEqual(_축척_텍스트_정리("NONE"), "NONE")

    def test_na_keyword(self):
        self.assertEqual(_축척_텍스트_정리("N/A"), "NONE")

    def test_spaces_in_ratio(self):
        self.assertEqual(_축척_텍스트_정리("1 : 50"), "1/50")

    def test_no_match_returns_x(self):
        self.assertEqual(_축척_텍스트_정리("도면명"), "X")

    def test_case_insensitive_none(self):
        self.assertEqual(_축척_텍스트_정리("none"), "NONE")

    def test_comma_in_denominator_stripped(self):
        self.assertEqual(_축척_텍스트_정리("1:1,000"), "1/1000")


# ============================================================================
# _extract_drawing_number
# ============================================================================
class TestExtractDrawingNumber(unittest.TestCase):
    def test_basic_extract(self):
        self.assertEqual(_extract_drawing_number("AA-001"), "AA-001")

    def test_returns_none_for_empty(self):
        self.assertIsNone(_extract_drawing_number(""))

    def test_a1_excluded(self):
        self.assertIsNone(_extract_drawing_number("A1"))

    def test_a3_excluded(self):
        self.assertIsNone(_extract_drawing_number("A3"))

    def test_a0_excluded(self):
        self.assertIsNone(_extract_drawing_number("A0"))

    def test_number_with_context(self):
        result = _extract_drawing_number("도면번호: BB-123-456")
        self.assertEqual(result, "BB-123-456")

    def test_prefix_ending_with_do_excluded(self):
        result = _extract_drawing_number("지하도-001")
        self.assertIsNone(result)

    def test_prefix_ending_with_pyo_excluded(self):
        result = _extract_drawing_number("일람표-001")
        self.assertIsNone(result)

    def test_multi_segment_number(self):
        result = _extract_drawing_number("AA-001-002-003")
        self.assertIsNotNone(result)
        self.assertIn("AA", result)

    def test_no_number_in_text(self):
        self.assertIsNone(_extract_drawing_number("평면도 입면도 단면도"))


# ============================================================================
# _extract_dong_from_title
# ============================================================================
class TestExtractDong(unittest.TestCase):
    def test_basic_dong(self):
        self.assertEqual(_extract_dong_from_title("1동 평면도"), "1동")

    def test_alpha_dong(self):
        self.assertEqual(_extract_dong_from_title("A동 입면도"), "A동")

    def test_excluded_word_공동(self):
        self.assertEqual(_extract_dong_from_title("공동주택 배치도"), "")

    def test_excluded_word_자동(self):
        self.assertEqual(_extract_dong_from_title("자동화 시스템 도면"), "")

    def test_excluded_word_수동(self):
        self.assertEqual(_extract_dong_from_title("수동 조작 패널"), "")

    def test_no_dong(self):
        self.assertEqual(_extract_dong_from_title("지하주차장 평면도"), "")

    def test_range_dong(self):
        result = _extract_dong_from_title("1동~3동 배치도")
        self.assertIn("동", result)

    def test_empty(self):
        self.assertEqual(_extract_dong_from_title(""), "")


# ============================================================================
# _clean_text_from_headers
# ============================================================================
class TestCleanFromHeaders(unittest.TestCase):
    def test_removes_drawing_no_header(self):
        result = _clean_text_from_headers("도면번호 AA-001")
        self.assertNotIn("도면번호", result)
        self.assertIn("AA-001", result)

    def test_removes_title_header(self):
        result = _clean_text_from_headers("TITLE 입면도")
        self.assertNotIn("TITLE", result)
        self.assertIn("입면도", result)

    def test_removes_scale_header(self):
        result = _clean_text_from_headers("SCALE 1:100")
        self.assertNotIn("SCALE", result)

    def test_empty_stays_empty(self):
        self.assertEqual(_clean_text_from_headers(""), "")

    def test_no_headers_unchanged(self):
        result = _clean_text_from_headers("일반 텍스트")
        self.assertEqual(result, "일반 텍스트")

    def test_case_insensitive(self):
        result = _clean_text_from_headers("title 입면도")
        self.assertNotIn("title", result.lower())


# ============================================================================
# _transform_xref_texts
# ============================================================================
class TestTransformXrefTexts(unittest.TestCase):
    def _approx_eq(self, a, b, tol=1e-6):
        return abs(a - b) < tol

    def test_identity_transform(self):
        texts = [(10.0, 20.0, "TEST", 5.0)]
        result = _transform_xref_texts(texts, 0.0, 0.0, 1.0, 1.0, 0.0)
        self.assertEqual(len(result), 1)
        rx, ry, txt, rh = result[0]
        self.assertTrue(self._approx_eq(rx, 10.0))
        self.assertTrue(self._approx_eq(ry, 20.0))
        self.assertEqual(txt, "TEST")
        self.assertTrue(self._approx_eq(rh, 5.0))

    def test_translation(self):
        texts = [(0.0, 0.0, "A", 1.0)]
        result = _transform_xref_texts(texts, 100.0, 200.0, 1.0, 1.0, 0.0)
        rx, ry, _, _ = result[0]
        self.assertTrue(self._approx_eq(rx, 100.0))
        self.assertTrue(self._approx_eq(ry, 200.0))

    def test_scaling(self):
        texts = [(1.0, 1.0, "A", 2.0)]
        result = _transform_xref_texts(texts, 0.0, 0.0, 2.0, 3.0, 0.0)
        rx, ry, _, rh = result[0]
        self.assertTrue(self._approx_eq(rx, 2.0))
        self.assertTrue(self._approx_eq(ry, 3.0))
        self.assertTrue(self._approx_eq(rh, 6.0))  # h * yscale

    def test_90_degree_rotation(self):
        texts = [(1.0, 0.0, "A", 1.0)]
        result = _transform_xref_texts(texts, 0.0, 0.0, 1.0, 1.0, 90.0)
        rx, ry, _, _ = result[0]
        # (1,0) rotated 90° → (0, 1)
        self.assertTrue(self._approx_eq(rx, 0.0, tol=1e-5))
        self.assertTrue(self._approx_eq(ry, 1.0, tol=1e-5))

    def test_180_degree_rotation(self):
        texts = [(1.0, 0.0, "A", 1.0)]
        result = _transform_xref_texts(texts, 0.0, 0.0, 1.0, 1.0, 180.0)
        rx, ry, _, _ = result[0]
        # (1,0) rotated 180° → (-1, 0)
        self.assertTrue(self._approx_eq(rx, -1.0, tol=1e-5))
        self.assertTrue(self._approx_eq(ry, 0.0, tol=1e-5))

    def test_empty_input(self):
        result = _transform_xref_texts([], 0.0, 0.0, 1.0, 1.0, 0.0)
        self.assertEqual(result, [])

    def test_text_and_height_preserved(self):
        texts = [(5.0, 5.0, "도면명", 10.0)]
        result = _transform_xref_texts(texts, 0.0, 0.0, 1.0, 1.0, 0.0)
        _, _, txt, h = result[0]
        self.assertEqual(txt, "도면명")
        self.assertTrue(self._approx_eq(h, 10.0))


# ============================================================================
# build_report
# ============================================================================
class TestBuildReport(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.out_path = os.path.join(self.tmpdir, "report.xlsx")

    def _make_list_df(self, rows):
        return pd.DataFrame(rows, columns=["도면번호(LIST)", "구분_LIST(동)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)"])

    def _make_dwg_df(self, rows):
        return pd.DataFrame(rows, columns=["파일명", "도면번호(DWG)", "구분_DWG(동)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])

    def test_both_empty_no_file_created(self):
        build_report(pd.DataFrame(), pd.DataFrame(), self.out_path)
        self.assertFalse(os.path.exists(self.out_path))

    def test_matching_rows_produce_excel(self):
        list_df = self._make_list_df([("AA-001", "", "1층 평면도", "1/100", "1/200")])
        dwg_df  = self._make_dwg_df([("A101.dwg", "AA-001", "", "1층 평면도", "1/100", "1/200")])
        build_report(list_df, dwg_df, self.out_path)
        self.assertTrue(os.path.exists(self.out_path))

    def test_status_일치(self):
        list_df = self._make_list_df([("AA-001", "", "평면도", "1/100", "X")])
        dwg_df  = self._make_dwg_df([("A101.dwg", "AA-001", "", "평면도", "1/100", "X")])
        build_report(list_df, dwg_df, self.out_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb.active
        header = {cell.value: cell.column for cell in ws[1]}
        status_col = header.get("상태")
        self.assertIsNotNone(status_col)
        self.assertEqual(ws.cell(2, status_col).value, "일치")

    def test_status_dwg_누락(self):
        list_df = self._make_list_df([("AA-999", "", "없는도면", "X", "X")])
        dwg_df  = self._make_dwg_df([])
        build_report(list_df, dwg_df, self.out_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb.active
        header = {cell.value: cell.column for cell in ws[1]}
        self.assertEqual(ws.cell(2, header["상태"]).value, "DWG 누락")

    def test_status_목록표_누락(self):
        list_df = self._make_list_df([])
        dwg_df  = self._make_dwg_df([("A999.dwg", "AA-999", "", "없는도면", "X", "X")])
        build_report(list_df, dwg_df, self.out_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb.active
        header = {cell.value: cell.column for cell in ws[1]}
        self.assertEqual(ws.cell(2, header["상태"]).value, "목록표 누락")

    def test_key_normalization(self):
        # 공백/하이픈 차이가 있어도 같은 번호로 매칭되어야 함
        list_df = self._make_list_df([("AA 001", "", "평면도", "X", "X")])
        dwg_df  = self._make_dwg_df([("A101.dwg", "AA-001", "", "평면도", "X", "X")])
        build_report(list_df, dwg_df, self.out_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb.active
        header = {cell.value: cell.column for cell in ws[1]}
        self.assertEqual(ws.cell(2, header["상태"]).value, "일치")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
